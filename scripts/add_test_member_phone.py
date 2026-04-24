"""Append test members (OTP phone list) to master Excel and optional Firestore.

The numbers are stored on the `cellPhone` column (XXX-XXX-XXXX) for Firebase Phone Auth testing.
Register the same E.164 numbers (+1...) under Firebase → Authentication → Phone numbers for
testing (Spark) or use Blaze for real SMS.

Usage (from project root):
  pip install -e ".[import]"
  python scripts/add_test_member_phone.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
import import_members_firestore as imf  # noqa: E402

import openpyxl

EXCEL = ROOT / "Master_List_Format_1500_Records.xlsx"
SHEET_NAME = "Master List Format"

# 10-digit US local numbers (no +1) — for testing; deduplicated. Display as XXX-XXX-XXXX in Excel.
TEST_PHONE_DIGITS: list[str] = list(
    dict.fromkeys(
        [
            "9176607617",
            "9734930439",
            "5512479193",
            "7853939651",
            "2135372528",
            "5512634892",
            "5513584363",
            "9146123602",
            "9174998739",
            "9175785306",
        ]
    )
)

# One row per test phone: unique sl / membership (avoid colliding with 999888 legacy test)
SL_BASE = 9971
MEMBERSHIP_BASE = 999901


def display_phone(digits: str) -> str:
    if len(digits) != 10 or not digits.isdigit():
        raise ValueError(f"expected 10 digits, got {digits!r}")
    return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"


def build_row(sl: int, membership_number: int, phone_display: str, test_index: int) -> tuple:
    email = f"test.otp{test_index:02d}@example.com"
    return (
        sl,
        "Test",
        f"Otp{test_index}",
        "",
        "LM",
        membership_number,
        "Test",
        1,
        2026,
        "2026-01-15",
        "1 Test Ln",
        "",
        "Edison",
        "NJ",
        "08820",
        "",  # homePhone
        "",  # businessPhone
        phone_display,  # cellPhone
        email,
        "",  # business
        "",  # alternatePhone
        "",  # childDetail1
        "",  # childDetail2
        "",  # childDetail3
        "",  # childDetail4
    )


def main() -> None:
    if len(imf.FIELD_NAMES) != 25:
        raise SystemExit("Expected 25 columns in FIELD_NAMES")

    if not EXCEL.is_file():
        print(f"Excel not found: {EXCEL}", file=sys.stderr)
        raise SystemExit(1)

    rows: list[tuple] = []
    for i, digits in enumerate(TEST_PHONE_DIGITS, start=1):
        row = build_row(
            SL_BASE + (i - 1),
            MEMBERSHIP_BASE + (i - 1),
            display_phone(digits),
            i,
        )
        if len(row) != 25:
            raise SystemExit("build_row() must have 25 values")
        rows.append(row)

    wb = openpyxl.load_workbook(EXCEL)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet {SHEET_NAME!r} not found: {wb.sheetnames}", file=sys.stderr)
        raise SystemExit(1)
    ws = wb[SHEET_NAME]
    mem_col = 1 + imf.FIELD_NAMES.index("membershipNumber")
    existing: set[int] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=mem_col).value
        if v is not None and str(v).strip().isdigit():
            existing.add(int(str(v).strip()))
    to_write = [
        r
        for r in rows
        if r[imf.FIELD_NAMES.index("membershipNumber")] not in existing
    ]
    if not to_write:
        print(
            f"Test rows (memberships {MEMBERSHIP_BASE}-{MEMBERSHIP_BASE + len(TEST_PHONE_DIGITS) - 1}) "
            f"already in {EXCEL.name}; nothing to append."
        )
    else:
        for row in to_write:
            new_row = ws.max_row + 1
            for col, _name in enumerate(imf.FIELD_NAMES, start=1):
                val = row[col - 1]
                ws.cell(row=new_row, column=col, value=val)
        wb.save(EXCEL)
        print(
            f"Appended {len(to_write)} row(s) to {EXCEL.name} "
            f"(memberships {MEMBERSHIP_BASE}-{MEMBERSHIP_BASE + len(TEST_PHONE_DIGITS) - 1})."
        )

    if not (ROOT / "serviceAccount.json").is_file() and not __import__("os").environ.get(
        "GOOGLE_APPLICATION_CREDENTIALS"
    ):
        print("No serviceAccount.json — Excel only. Add credentials to write Firestore.")
        return

    db = imf.init_firebase()
    e164: list[str] = []
    for row in rows:
        member = imf.row_to_member(row)
        member["sourceExcelPath"] = str(EXCEL)
        mnum = member.get("membershipNumber")
        if mnum is None:
            continue
        ref = db.collection("members").document(str(mnum))
        ref.set(member, merge=True)
        d = re.sub(r"\D", "", str(row[imf.FIELD_NAMES.index("cellPhone")]))
        e164.append(f"+1{d}")
    last = MEMBERSHIP_BASE + len(TEST_PHONE_DIGITS) - 1
    print(
        f"Firestore: set members/{MEMBERSHIP_BASE}-{last} (merge), cells "
        f"{', '.join(e164)}"
    )
    n = len(TEST_PHONE_DIGITS)
    print(f"Search: Otp1..Otp{n}, Test, or the 10-digit cell values.")


if __name__ == "__main__":
    main()
