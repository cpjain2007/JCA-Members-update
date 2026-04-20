"""Generate a synthetic Master List Format sheet (1500 rows) matching the template workbook."""

from __future__ import annotations

import random
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "20250420_LifeMember List Update_Format.xlsx"
OUTPUT = ROOT / "Master_List_Format_1500_Records.xlsx"
SHEET_NAME = "Master List Format"
N_ROWS = 1500

FIRST_NAMES = [
    "Aarav", "Priya", "Rohan", "Neha", "Vikram", "Anika", "Kiran", "Sneha", "Dev", "Isha",
    "Arjun", "Meera", "Sanjay", "Pooja", "Rahul", "Kavya", "Aditya", "Divya", "Nikhil", "Riya",
    "James", "Emily", "Michael", "Sarah", "David", "Jessica", "Daniel", "Ashley", "Matthew", "Amanda",
]
LAST_NAMES = [
    "Shah", "Patel", "Mehta", "Kapoor", "Singh", "Kumar", "Reddy", "Agarwal", "Malhotra", "Bansal",
    "Joshi", "Desai", "Iyer", "Nair", "Gupta", "Verma", "Chopra", "Sen", "Rao", "Khanna",
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Martinez", "Lee",
]
CITIES = [
    ("Jersey City", "NJ", "07306"),
    ("Jersey City", "NJ", "07302"),
    ("Newark", "NJ", "07102"),
    ("Edison", "NJ", "08817"),
    ("Princeton", "NJ", "08540"),
    ("Forest Hills", "NY", "11375"),
    ("Flushing", "NY", "11354"),
    ("Brooklyn", "NY", "11201"),
    ("Manhattan", "NY", "10001"),
    ("Stamford", "CT", "06901"),
]
STATUS_CHOICES = [None, "JCA office", "JCA office", "Active", "Pending"]
BUSINESS_CHOICES = [None, "IT", "Finance", "Healthcare", "Education", "Retail", "Engineering", "Legal"]


def random_edited_date(rng: random.Random) -> datetime:
    start = datetime(2008, 1, 1)
    end = datetime(2026, 4, 19)
    days = (end - start).days
    return (start + timedelta(days=rng.randint(0, days))).replace(
        hour=0, minute=0, second=0, microsecond=0
    )


def phone(rng: random.Random) -> str:
    return f"{rng.randint(200, 999)}-{rng.randint(200, 999)}-{rng.randint(1000, 9999)}"


def build_rows(rng: random.Random, membership_start: int) -> list[tuple]:
    rows: list[tuple] = []
    used_membership = set()
    for i in range(1, N_ROWS + 1):
        last = rng.choice(LAST_NAMES)
        first = rng.choice(FIRST_NAMES)
        spouse = rng.choice(FIRST_NAMES)
        m = membership_start + i
        while m in used_membership:
            m += 1
        used_membership.add(m)
        city, st, zip_str = rng.choice(CITIES)
        zip_val = int(zip_str)
        edited = random_edited_date(rng)
        year = rng.randint(2005, 2026)
        receipt = rng.randint(10000, 99999)
        apt = None if rng.random() > 0.35 else str(rng.randint(1, 50))
        home = phone(rng)
        biz_phone = None if rng.random() > 0.55 else phone(rng)
        cell = None if rng.random() > 0.45 else phone(rng)
        email_local = f"{first.lower()}.{last.lower()}{i}"
        email = f"{email_local}@example.com"
        business = rng.choice(BUSINESS_CHOICES)
        col21 = None if rng.random() > 0.6 else phone(rng)
        children = (None, None, None, None)
        if rng.random() > 0.85:
            children = (
                f"{rng.choice(FIRST_NAMES)} ({rng.randint(1, 18)}y)",
                None,
                None,
                None,
            )
        addr_num = rng.randint(1, 9999)
        addr = f"{addr_num} {rng.choice(['Oak', 'Maple', 'Park', 'Lake', 'River', 'Hill'])} {rng.choice(['St', 'Ave', 'Rd', 'Blvd'])}"
        status = rng.choice(STATUS_CHOICES)
        rows.append(
            (
                i,
                last,
                first,
                spouse,
                "LM",
                m,
                status,
                receipt,
                year,
                edited,
                addr,
                apt,
                city,
                st,
                zip_val,
                home,
                biz_phone,
                cell,
                email,
                business,
                col21,
                *children,
            )
        )
    return rows


def main() -> None:
    rng = random.Random(42)
    tpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
    src = tpl[SHEET_NAME]
    formats = [src.cell(2, c).number_format for c in range(1, 26)]

    out = openpyxl.Workbook()
    out.remove(out.active)
    dst = out.create_sheet(SHEET_NAME)

    for c in range(1, 26):
        sc, dc = src.cell(1, c), dst.cell(1, c)
        dc.value = sc.value
        _copy_cell_style(sc, dc)

    for c in range(1, 26):
        col_letter = get_column_letter(c)
        dim = src.column_dimensions.get(col_letter)
        if dim and dim.width:
            dst.column_dimensions[col_letter].width = dim.width

    membership_start = 50_000
    for row_idx, row in enumerate(build_rows(rng, membership_start), start=2):
        for c, val in enumerate(row, start=1):
            cell = dst.cell(row=row_idx, column=c, value=val)
            cell.number_format = formats[c - 1]

    out.save(OUTPUT)
    tpl.close()
    print(f"Wrote {OUTPUT} ({N_ROWS} data rows).")


if __name__ == "__main__":
    main()
